import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df_ec = pd.read_excel("E-Ticaret_MMM_v2.xlsx")
df_fg = pd.read_excel("FMCG_MMM_v2.xlsx")

def analyser_manquants(df, nom_dataset):
    resultats = []
    for col in df.columns:
        n_manq = df[col].isnull().sum()
        pct = round(n_manq / len(df) * 100, 2)
        dtype = str(df[col].dtype)
        if "float" in dtype or "int" in dtype:
            type_var = "Numérique"
        elif "datetime" in dtype:
            type_var = "Date"
        else:
            type_var = "Catégorielle"

        if n_manq == 0:
            traitement = "Aucun traitement requis"
        elif "Retard" in col:
            traitement = "Valeur manquante structurelle (décalage temporel)"
        elif type_var == "Numérique":
            traitement = "Exclusion de l'observation"
        else:
            traitement = "Exclusion de l'observation"

        resultats.append({
            "Jeu de données": nom_dataset,
            "Variable": col.strip(),
            "Type": type_var,
            "Observations totales": len(df),
            "Valeurs manquantes": int(n_manq),
            "Pourcentage (%)": pct,
            "Traitement appliqué": traitement,
        })
    return resultats

rows_ec = analyser_manquants(df_ec, "E-Commerce")
rows_fg = analyser_manquants(df_fg, "FMCG")

all_rows = rows_ec + rows_fg
df_out = pd.DataFrame(all_rows)
df_manq = df_out[df_out["Valeurs manquantes"] > 0].reset_index(drop=True)
df_ok = df_out[df_out["Valeurs manquantes"] == 0].reset_index(drop=True)

wb = Workbook()

def ecrire_feuille(ws, titre_tab, df_data, couleur_entete="1F3864"):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_cols = len(df_data.columns)
    last_col = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = titre_tab
    ws["A1"].font = Font(name="Times New Roman", bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for col_idx, col_name in enumerate(df_data.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(name="Times New Roman", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=couleur_entete)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 28

    couleur_paire = "DCE6F1"
    for i, row_data in df_data.iterrows():
        row_excel = i + 3
        fill_color = couleur_paire if i % 2 == 0 else "FFFFFF"
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_excel, column=col_idx, value=val)
            c.font = Font(name="Times New Roman", size=10)
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[row_excel].height = 15

    note_row = len(df_data) + 4
    ws.cell(row=note_row, column=1, value="Source : Auteur, à partir des données propriétaires E-Commerce et FMCG (2022–2024).")
    ws.cell(row=note_row, column=1).font = Font(name="Times New Roman", italic=True, size=9)

ws1 = wb.active
ws1.title = "Valeurs manquantes"
ecrire_feuille(
    ws1,
    "Tableau 2a — Variables présentant des valeurs manquantes",
    df_manq,
    couleur_entete="C00000"
)
ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 36
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 20
ws1.column_dimensions["E"].width = 20
ws1.column_dimensions["F"].width = 18
ws1.column_dimensions["G"].width = 42

ws2 = wb.create_sheet("Résumé global")
resume = pd.DataFrame([
    {
        "Jeu de données": "E-Commerce",
        "Observations totales": len(df_ec),
        "Variables totales": df_ec.shape[1],
        "Variables sans valeur manquante": int((df_ec.isnull().sum() == 0).sum()),
        "Variables avec valeur manquante": int((df_ec.isnull().sum() > 0).sum()),
        "Total valeurs manquantes": int(df_ec.isnull().sum().sum()),
        "Taux global de complétude (%)": round(100 - df_ec.isnull().sum().sum() / (df_ec.shape[0] * df_ec.shape[1]) * 100, 2),
    },
    {
        "Jeu de données": "FMCG",
        "Observations totales": len(df_fg),
        "Variables totales": df_fg.shape[1],
        "Variables sans valeur manquante": int((df_fg.isnull().sum() == 0).sum()),
        "Variables avec valeur manquante": int((df_fg.isnull().sum() > 0).sum()),
        "Total valeurs manquantes": int(df_fg.isnull().sum().sum()),
        "Taux global de complétude (%)": round(100 - df_fg.isnull().sum().sum() / (df_fg.shape[0] * df_fg.shape[1]) * 100, 2),
    }
])
ecrire_feuille(
    ws2,
    "Tableau 2b — Résumé global du traitement des valeurs manquantes",
    resume,
    couleur_entete="1F3864"
)
for col_letter in ["A","B","C","D","E","F","G"]:
    ws2.column_dimensions[col_letter].width = 28

output_path = "Tableau_02_Valeurs_Manquantes.xlsx"
wb.save(output_path)
print(f"Fichier sauvegardé : {output_path}")
