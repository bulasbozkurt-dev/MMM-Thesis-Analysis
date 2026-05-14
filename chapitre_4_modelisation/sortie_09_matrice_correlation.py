import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df_ec = pd.read_excel("E-Ticaret_MMM_v2.xlsx")
df_fg = pd.read_excel("FMCG_MMM_v2.xlsx")
df_ec.columns = [c.strip() for c in df_ec.columns]
df_fg.columns = [c.strip() for c in df_fg.columns]

VARS_EC = [
    "D : Commandes (canal)",
    "M : Google",
    "M : Meta",
    "M : Influenceur",
    "M : Indice de dépenses",
    "M : Retard t-1",
    "M : Retard t-7",
    "D : Retard t-1",
    "Choc : Taux de change",
    "Choc : Crise logistique",
    "Facteur saisonnier",
    "Effet du dimanche",
    "Dépenses des concurrents",
    "Part de voix (%)",
    "Taux de conversion (%)",
]

VARS_FG = [
    "D : Volume des ventes",
    "M : Dépenses marketing totales",
    "M : Pondéré",
    "M : Retard 1 semaine",
    "M : Retard 4 semaines",
    "D : Retard 1 semaine",
    "Choc : Taux de change",
    "Choc : Impact du séisme",
    "Facteur saisonnier",
    "Dépenses marketing des concurrents",
    "Part de voix (%)",
    "Part de marché (%)",
    "Rupture de stock (%)",
    "Stock (jours)",
    "Efficacité marketing",
]

LABELS_EC = [
    "D : Commandes",
    "M : Google",
    "M : Meta",
    "M : Influenceur",
    "M : Indice dép.",
    "M : Retard t−1",
    "M : Retard t−7",
    "D : Retard t−1",
    "Choc : Tx change",
    "Choc : Logist.",
    "Fact. saisonnier",
    "Effet dimanche",
    "Dép. concurrents",
    "Part de voix",
    "Tx conversion",
]

LABELS_FG = [
    "D : Volume ventes",
    "M : Dép. totales",
    "M : Pondéré",
    "M : Retard 1 sem.",
    "M : Retard 4 sem.",
    "D : Retard 1 sem.",
    "Choc : Tx change",
    "Choc : Séisme",
    "Fact. saisonnier",
    "Dép. concurrents",
    "Part de voix",
    "Part de marché",
    "Rupture stock",
    "Stock (jours)",
    "Efficacité mktg",
]

def corr_matrix(df, vars_list):
    sub = df[[v for v in vars_list if v in df.columns]].dropna()
    return sub.corr()

corr_ec = corr_matrix(df_ec, VARS_EC)
corr_fg = corr_matrix(df_fg, VARS_FG)

def hex_gradient(val):
    if pd.isna(val):
        return "FFFFFF"
    if val == 1.0:
        return "D9D9D9"
    if val > 0:
        intensity = int(255 - val * 180)
        return f"{intensity:02X}FF{intensity:02X}"
    else:
        intensity = int(255 - abs(val) * 180)
        return f"FF{intensity:02X}{intensity:02X}"

thin  = Side(style="thin",   color="BBBBBB")
thick = Side(style="medium", color="1F3864")
b = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def ecrire_matrice(wb, corr_df, labels, titre_tab, nom_feuille, premier=True):
    ws = wb.active if premier else wb.create_sheet(nom_feuille)
    if not premier:
        pass
    else:
        ws.title = nom_feuille

    n = len(labels)
    last_col = get_column_letter(n + 1)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = titre_tab
    ws["A1"].font      = Font(name="Times New Roman", bold=True, size=11)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.cell(row=2, column=1, value="Variable").font = Font(
        name="Times New Roman", bold=True, color="FFFFFF", size=9)
    ws.cell(row=2, column=1).fill = PatternFill("solid", fgColor="1F3864")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=1).border = b

    for ci, lbl in enumerate(labels, 2):
        c = ws.cell(row=2, column=ci, value=lbl)
        c.font      = Font(name="Times New Roman", bold=True, color="FFFFFF", size=8)
        c.fill      = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True, text_rotation=45)
        c.border    = b
    ws.row_dimensions[2].height = 70

    for ri, lbl_row in enumerate(labels, 3):
        c = ws.cell(row=ri, column=1, value=lbl_row)
        c.font      = Font(name="Times New Roman", bold=True, size=8)
        c.fill      = PatternFill("solid", fgColor="DCE6F1")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = b
        ws.row_dimensions[ri].height = 16

        col_name = corr_df.columns[ri - 3] if ri - 3 < len(corr_df.columns) else None
        for ci, lbl_col in enumerate(labels, 2):
            row_name = corr_df.index[ri - 3] if ri - 3 < len(corr_df.index) else None
            col_name2 = corr_df.columns[ci - 2] if ci - 2 < len(corr_df.columns) else None
            try:
                val = corr_df.loc[row_name, col_name2]
            except Exception:
                val = np.nan

            cell = ws.cell(row=ri, column=ci)
            if pd.isna(val):
                cell.value = "—"
                cell.fill  = PatternFill("solid", fgColor="EEEEEE")
            elif ci - 2 < ri - 3:
                cell.value = round(val, 3)
                cell.fill  = PatternFill("solid", fgColor=hex_gradient(val))
                cell.font  = Font(name="Times New Roman", size=8,
                                  bold=abs(val) >= 0.5)
            elif ci - 2 == ri - 3:
                cell.value = "1"
                cell.fill  = PatternFill("solid", fgColor="D9D9D9")
                cell.font  = Font(name="Times New Roman", size=8, bold=True)
            else:
                cell.value = ""
                cell.fill  = PatternFill("solid", fgColor="F5F5F5")

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = b

    ws.column_dimensions["A"].width = 20
    for ci in range(2, n + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 8

    note_row = n + 4
    ws.cell(row=note_row, column=1,
            value="Note : Valeurs en gras = |r| ≥ 0,50. "
                  "Vert = corrélation positive. Rouge = corrélation négative. "
                  "Matrice triangulaire inférieure.")
    ws.cell(row=note_row, column=1).font = Font(name="Times New Roman",
                                                italic=True, size=8.5)
    ws.cell(row=note_row+1, column=1,
            value="Source : Auteur, à partir des données propriétaires (2022–2024).")
    ws.cell(row=note_row+1, column=1).font = Font(name="Times New Roman",
                                                   italic=True, size=8.5)

ecrire_matrice(
    wb, corr_ec, LABELS_EC,
    "Tableau 9a — Matrice de corrélation — E-Commerce",
    "Corrélations E-Commerce", premier=True
)
ecrire_matrice(
    wb, corr_fg, LABELS_FG,
    "Tableau 9b — Matrice de corrélation — FMCG",
    "Corrélations FMCG", premier=False
)

output_path = "Tableau_09_Matrice_Correlation.xlsx"
wb.save(output_path)
print(f"Fichier sauvegardé : {output_path}")
