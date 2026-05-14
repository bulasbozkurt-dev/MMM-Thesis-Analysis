import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df_ec = pd.read_excel("E-Ticaret_MMM_v2.xlsx")
df_fg = pd.read_excel("FMCG_MMM_v2.xlsx")

cols_ec = [c.strip() for c in df_ec.columns]
cols_fg = [c.strip() for c in df_fg.columns]

def categoriser(col):
    if col.startswith("M :") or "Dépenses marketing" in col or "Pondéré" in col or "Poids" in col:
        return "Variable explicative (Marketing)"
    elif col.startswith("D :"):
        return "Variable dépendante"
    elif col.startswith("ε :"):
        return "Variable dépendante (variation)"
    elif col.startswith("Choc :"):
        return "Variable de choc"
    elif col.startswith("C :"):
        return "Variable de coût"
    elif col.startswith("R :"):
        return "Variable de revenu"
    elif col.startswith("π :"):
        return "Variable de profit"
    elif col in ["Facteur saisonnier", "Effet du dimanche", "Week-end", "Fin de mois", "Trimestre"]:
        return "Effet fixe / Saisonnalité"
    elif col in ["Date", "Année", "Mois", "Numéro de semaine", "Jour", "Trimestre"]:
        return "Variable temporelle"
    elif col in ["Catégorie", "SKU", "Canal", "Région", "Segment"]:
        return "Variable de segmentation"
    elif "Retard" in col:
        return "Variable de retard"
    elif "concurrent" in col.lower() or "Part de voix" in col or "Part de marché" in col:
        return "Variable de contrôle (concurrence)"
    elif "Taux de conversion" in col or "Délai" in col or "Score" in col or "Pression" in col:
        return "Variable de contrôle (opérationnel)"
    elif "Rupture" in col or "Stock" in col or "Risque" in col or "Perte" in col or "Efficacité" in col:
        return "Variable de contrôle (stock)"
    elif "Prix" in col or "Coût" in col or "Ratio" in col:
        return "Variable de prix / coût unitaire"
    else:
        return "Autre"

def notation(col):
    if col.startswith("M :"):
        return "M"
    elif col.startswith("D :") or col.startswith("ε :"):
        return "D / ε"
    elif col.startswith("Choc :"):
        return "δ"
    elif col.startswith("C :"):
        return "C"
    elif col.startswith("R :"):
        return "R"
    elif col.startswith("π :"):
        return "π"
    elif "Retard" in col:
        return "t−k"
    elif col in ["Facteur saisonnier", "Effet du dimanche", "Week-end", "Fin de mois"]:
        return "γ"
    else:
        return "—"

rows = []
all_cols = sorted(set(cols_ec + cols_fg))

for col in all_cols:
    dans_ec = "✓" if col in cols_ec else "—"
    dans_fg = "✓" if col in cols_fg else "—"
    cat = categoriser(col)
    nota = notation(col)
    rows.append({
        "Variable (nom original)": col,
        "Catégorie": cat,
        "Notation MCO": nota,
        "E-Commerce": dans_ec,
        "FMCG": dans_fg,
    })

df_corr = pd.DataFrame(rows)
ordre_cat = [
    "Variable dépendante",
    "Variable dépendante (variation)",
    "Variable explicative (Marketing)",
    "Variable de retard",
    "Variable de choc",
    "Effet fixe / Saisonnalité",
    "Variable de contrôle (concurrence)",
    "Variable de contrôle (opérationnel)",
    "Variable de contrôle (stock)",
    "Variable temporelle",
    "Variable de segmentation",
    "Variable de revenu",
    "Variable de coût",
    "Variable de prix / coût unitaire",
    "Variable de profit",
    "Autre",
]
df_corr["_ordre"] = df_corr["Catégorie"].apply(lambda x: ordre_cat.index(x) if x in ordre_cat else 99)
df_corr = df_corr.sort_values(["_ordre", "Variable (nom original)"]).drop(columns="_ordre").reset_index(drop=True)

wb = Workbook()
ws = wb.active
ws.title = "Correspondance des variables"

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

titre = "Tableau 3 — Correspondance des variables entre les deux jeux de données"
ws.merge_cells("A1:E1")
ws["A1"] = titre
ws["A1"].font = Font(name="Times New Roman", bold=True, size=12)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 22

entetes = list(df_corr.columns)
couleur_entete = "1F3864"
for col_idx, val in enumerate(entetes, start=1):
    cell = ws.cell(row=2, column=col_idx, value=val)
    cell.font = Font(name="Times New Roman", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=couleur_entete)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[2].height = 28

couleurs_cat = {
    "Variable dépendante": "F4CCCC",
    "Variable dépendante (variation)": "F4CCCC",
    "Variable explicative (Marketing)": "D9EAD3",
    "Variable de retard": "EAD1DC",
    "Variable de choc": "FCE5CD",
    "Effet fixe / Saisonnalité": "FFF2CC",
    "Variable de contrôle (concurrence)": "CFE2F3",
    "Variable de contrôle (opérationnel)": "CFE2F3",
    "Variable de contrôle (stock)": "CFE2F3",
    "Variable temporelle": "EFEFEF",
    "Variable de segmentation": "EFEFEF",
    "Variable de revenu": "D9D2E9",
    "Variable de coût": "D9D2E9",
    "Variable de prix / coût unitaire": "D9D2E9",
    "Variable de profit": "D9D2E9",
    "Autre": "FFFFFF",
}

for i, row_data in df_corr.iterrows():
    row_excel = i + 3
    cat = row_data["Catégorie"]
    fill_color = couleurs_cat.get(cat, "FFFFFF")
    for col_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=row_excel, column=col_idx, value=val)
        c.font = Font(name="Times New Roman", size=9)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        c.border = border
    ws.row_dimensions[row_excel].height = 14

ws.column_dimensions["A"].width = 44
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 10

note_row = len(df_corr) + 4
ws.cell(row=note_row, column=1, value="Note : ✓ = variable présente dans le jeu de données ; — = variable absente.")
ws.cell(row=note_row, column=1).font = Font(name="Times New Roman", italic=True, size=9)
ws.cell(row=note_row + 1, column=1, value="Source : Auteur, à partir des données propriétaires E-Commerce et FMCG (2022–2024).")
ws.cell(row=note_row + 1, column=1).font = Font(name="Times New Roman", italic=True, size=9)

output_path = "Tableau_03_Correspondance_Variables.xlsx"
wb.save(output_path)
print(f"Fichier sauvegardé : {output_path}")
