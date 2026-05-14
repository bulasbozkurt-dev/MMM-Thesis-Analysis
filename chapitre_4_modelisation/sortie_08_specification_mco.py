import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df_ec = pd.read_excel("E-Ticaret_MMM_v2.xlsx")
df_fg = pd.read_excel("FMCG_MMM_v2.xlsx")
df_ec.columns = [c.strip() for c in df_ec.columns]
df_fg.columns = [c.strip() for c in df_fg.columns]

SPEC = [
    {
        "section": "Spécification générale",
        "lignes": [
            ("Méthode d'estimation", "Moindres Carrés Ordinaires (MCO)", "Moindres Carrés Ordinaires (MCO)"),
            ("Variable dépendante", "D : Commandes (canal)", "D : Volume des ventes"),
            ("Transformation de la variable dépendante", "Aucune (niveau)", "Aucune (niveau)"),
            ("Fréquence temporelle", "Quotidienne", "Hebdomadaire"),
            ("Période d'estimation", "2022–2024", "2022–2024"),
            ("Nombre d'observations", str(df_ec.shape[0]), str(df_fg.shape[0])),
            ("Unités d'analyse", f"{df_ec['Canal'].nunique()} canaux × {df_ec['SKU'].nunique()} SKU",
             f"{df_fg['Région'].nunique()} régions × {df_fg['SKU'].nunique()} SKU"),
        ]
    },
    {
        "section": "Variables explicatives incluses",
        "lignes": [
            ("Variables marketing (M :)", "M : Google, M : Meta, M : Influenceur, M : Indice de dépenses",
             "M : Dépenses marketing totales, M : Pondéré"),
            ("Variables de retard", "M : Retard t−1, M : Retard t−7, D : Retard t−1",
             "M : Retard 1 semaine, M : Retard 4 semaines, D : Retard 1 semaine"),
            ("Variables de choc", "Choc : Taux de change, Choc : Crise logistique",
             "Choc : Taux de change, Choc : Impact du séisme"),
            ("Effets fixes / Saisonnalité", "Facteur saisonnier, Effet du dimanche, Week-end, Fin de mois",
             "Facteur saisonnier, Trimestre"),
            ("Variables de contrôle (concurrence)", "Dépenses des concurrents, Part de voix (%)",
             "Dépenses marketing des concurrents, Part de voix (%), Part de marché (%)"),
            ("Variables de contrôle (opérationnel)", "Taux de conversion (%), Délai de traitement, Score normalisé du délai",
             "Rupture de stock (%), Stock (jours), Efficacité marketing"),
        ]
    },
    {
        "section": "Hypothèses du modèle MCO",
        "lignes": [
            ("H1 — Linéarité", "Relation linéaire entre D et les variables explicatives", "Idem"),
            ("H2 — Exogénéité", "E(ε | X) = 0 ; absence de corrélation entre ε et les régresseurs", "Idem"),
            ("H3 — Homoscédasticité", "Var(ε | X) = σ² constant (à vérifier : test de Breusch-Pagan)", "Idem"),
            ("H4 — Absence d'autocorrélation", "Cov(εt, εs) = 0 pour t ≠ s (à vérifier : test de Durbin-Watson)", "Idem"),
            ("H5 — Absence de multicolinéarité parfaite", "Les variables explicatives ne sont pas parfaitement colinéaires (à vérifier : VIF)", "Idem"),
            ("H6 — Normalité des résidus", "ε ~ N(0, σ²) (à vérifier : test de Jarque-Bera)", "Idem"),
        ]
    },
    {
        "section": "Notation formelle du modèle",
        "lignes": [
            ("Équation générale",
             "Dt = α + Σk βk Mk,t + Σj γj Fj,t + Σm δm Sm,t + Σl λl Cl,t + εt",
             "Dt = α + Σk βk Mk,t + Σj γj Fj,t + Σm δm Sm,t + Σl λl Cl,t + εt"),
            ("α", "Constante du modèle (intercepte)", "Idem"),
            ("βk", "Coefficient de la variable marketing k", "Idem"),
            ("γj", "Coefficient de l'effet fixe / saisonnier j", "Idem"),
            ("δm", "Coefficient du choc m", "Idem"),
            ("λl", "Coefficient de la variable de contrôle l", "Idem"),
            ("εt", "Terme d'erreur à la période t", "Idem"),
        ]
    },
    {
        "section": "Critères d'évaluation prévus",
        "lignes": [
            ("Qualité d'ajustement", "R² ajusté", "R² ajusté"),
            ("Significativité globale", "Test F de Fisher (p < 0,05)", "Test F de Fisher (p < 0,05)"),
            ("Significativité individuelle", "Test t de Student (p < 0,05 ; p < 0,10)", "Idem"),
            ("Multicolinéarité", "Facteur d'inflation de la variance (VIF < 10)", "Idem"),
            ("Autocorrélation des résidus", "Statistique de Durbin-Watson (1,5 < DW < 2,5)", "Idem"),
            ("Hétéroscédasticité", "Test de Breusch-Pagan", "Idem"),
            ("Normalité des résidus", "Test de Jarque-Bera", "Idem"),
        ]
    },
]

wb = Workbook()
ws = wb.active
ws.title = "Spécification MCO"

thin  = Side(style="thin",   color="AAAAAA")
thick = Side(style="medium", color="1F3864")
b_data  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
b_sect  = Border(left=thick, right=thick, top=thick, bottom=thick)

COLS = ["Élément de spécification", "E-Commerce", "FMCG"]
N    = len(COLS)
last = get_column_letter(N)

ws.merge_cells(f"A1:{last}1")
ws["A1"] = "Tableau 8 — Résumé de la spécification du modèle MCO"
ws["A1"].font      = Font(name="Times New Roman", bold=True, size=12)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 22

for ci, h in enumerate(COLS, 1):
    c = ws.cell(row=2, column=ci, value=h)
    c.font      = Font(name="Times New Roman", bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = b_data
ws.row_dimensions[2].height = 24

SECT_COLORS = {
    "Spécification générale":           "D9EAD3",
    "Variables explicatives incluses":  "CFE2F3",
    "Hypothèses du modèle MCO":         "FFF2CC",
    "Notation formelle du modèle":      "EAD1DC",
    "Critères d'évaluation prévus":     "F4CCCC",
}

cur = 3
for bloc in SPEC:
    sec = bloc["section"]
    ws.merge_cells(f"A{cur}:{last}{cur}")
    c = ws.cell(row=cur, column=1, value=sec)
    c.font      = Font(name="Times New Roman", bold=True, size=10, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="2E5090")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = b_sect
    ws.row_dimensions[cur].height = 17
    cur += 1

    fill_hex = SECT_COLORS.get(sec, "FFFFFF")
    for i, (elem, val_ec, val_fg) in enumerate(bloc["lignes"]):
        bg = fill_hex if i % 2 == 0 else "FFFFFF"
        vals = [elem, val_ec, val_fg]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=cur, column=ci, value=v)
            c.font      = Font(name="Times New Roman", size=9)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            c.border = b_data
        ws.row_dimensions[cur].height = 28
        cur += 1

note = cur + 1
ws.cell(row=note, column=1,
        value="Note : MCO = Moindres Carrés Ordinaires. "
              "VIF = Facteur d'inflation de la variance. "
              "DW = Durbin-Watson. p = probabilité critique.")
ws.cell(row=note, column=1).font = Font(name="Times New Roman", italic=True, size=8.5)
ws.cell(row=note+1, column=1,
        value="Source : Auteur, à partir des données propriétaires E-Commerce et FMCG (2022–2024).")
ws.cell(row=note+1, column=1).font = Font(name="Times New Roman", italic=True, size=8.5)

ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 36

output_path = "Tableau_08_Specification_MCO.xlsx"
wb.save(output_path)
print(f"Fichier sauvegardé : {output_path}")
