import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy import stats

df_ec = pd.read_excel("E-Ticaret_MMM_v2.xlsx")
df_fg = pd.read_excel("FMCG_MMM_v2.xlsx")

df_ec.columns = [c.strip() for c in df_ec.columns]
df_fg.columns = [c.strip() for c in df_fg.columns]

GROUPES = {
    "Variable dépendante": {
        "ec": ["D : Commandes (canal)", "D : Total quotidien"],
        "fg": ["D : Volume des ventes"],
    },
    "Variables marketing": {
        "ec": ["M : Google", "M : Meta", "M : Influenceur", "M : Indice de dépenses"],
        "fg": ["M : Dépenses marketing totales", "M : Pondéré"],
    },
    "Variables de retard": {
        "ec": ["M : Retard t-1", "M : Retard t-7", "D : Retard t-1"],
        "fg": ["M : Retard 1 semaine", "M : Retard 4 semaines", "D : Retard 1 semaine"],
    },
    "Variables de choc": {
        "ec": ["Choc : Taux de change", "Choc : Crise logistique"],
        "fg": ["Choc : Taux de change", "Choc : Impact du séisme"],
    },
    "Effets fixes / Saisonnalité": {
        "ec": ["Facteur saisonnier", "Effet du dimanche"],
        "fg": ["Facteur saisonnier"],
    },
    "Variables de contrôle": {
        "ec": ["Dépenses des concurrents", "Part de voix (%)", "Taux de conversion (%)"],
        "fg": ["Dépenses marketing des concurrents", "Part de voix (%)", "Part de marché (%)"],
    },
    "Variables de revenu / profit": {
        "ec": ["R : Revenu (₺)", "π : Profit net (₺)", "π : Marge nette (%)"],
        "fg": ["R : Revenu (₺)", "π : Profit net (₺)", "π : Marge nette (%)"],
    },
}

def stats_col(df, col):
    s = df[col].dropna()
    if len(s) == 0:
        return ["—"] * 8
    sk = stats.skew(s)
    ku = stats.kurtosis(s)
    return [
        round(s.mean(), 2),
        round(s.median(), 2),
        round(s.std(), 2),
        round(s.min(), 2),
        round(s.max(), 2),
        int(s.count()),
        round(sk, 3),
        round(ku, 3),
    ]

rows = []
for groupe, cols_dict in GROUPES.items():
    ec_cols = cols_dict.get("ec", [])
    fg_cols = cols_dict.get("fg", [])
    all_vars = list(dict.fromkeys(ec_cols + fg_cols))
    rows.append({"__groupe": groupe, "__header": True})
    for var in all_vars:
        in_ec = var in df_ec.columns
        in_fg = var in df_fg.columns
        if in_ec:
            s = stats_col(df_ec, var)
            dataset = "E-Commerce"
        elif in_fg:
            s = stats_col(df_fg, var)
            dataset = "FMCG"
        else:
            continue
        rows.append({
            "__groupe": groupe,
            "__header": False,
            "Variable": var,
            "Jeu de données": dataset,
            "Moyenne": s[0],
            "Médiane": s[1],
            "Écart-type": s[2],
            "Minimum": s[3],
            "Maximum": s[4],
            "N": s[5],
            "Asymétrie": s[6],
            "Aplatissement": s[7],
        })
        if in_ec and in_fg:
            s2 = stats_col(df_fg, var)
            rows.append({
                "__groupe": groupe,
                "__header": False,
                "Variable": var,
                "Jeu de données": "FMCG",
                "Moyenne": s2[0],
                "Médiane": s2[1],
                "Écart-type": s2[2],
                "Minimum": s2[3],
                "Maximum": s2[4],
                "N": s2[5],
                "Asymétrie": s2[6],
                "Aplatissement": s2[7],
            })

wb = Workbook()
ws = wb.active
ws.title = "Statistiques descriptives"

thin = Side(style="thin", color="AAAAAA")
thick = Side(style="medium", color="1F3864")
border_data = Border(left=thin, right=thin, top=thin, bottom=thin)
border_group = Border(left=thick, right=thick, top=thick, bottom=thin)

ENTETES = ["Variable", "Jeu de données", "Moyenne", "Médiane",
           "Écart-type", "Minimum", "Maximum", "N", "Asymétrie", "Aplatissement"]
N_COLS = len(ENTETES)
last_col = get_column_letter(N_COLS)

ws.merge_cells(f"A1:{last_col}1")
ws["A1"] = "Tableau 5 — Statistiques descriptives des variables du modèle MCO"
ws["A1"].font = Font(name="Times New Roman", bold=True, size=12)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 22

for ci, h in enumerate(ENTETES, 1):
    c = ws.cell(row=2, column=ci, value=h)
    c.font = Font(name="Times New Roman", bold=True, color="FFFFFF", size=9.5)
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.row_dimensions[2].height = 26

GROUPE_COLORS = {
    "Variable dépendante":          "F4CCCC",
    "Variables marketing":          "D9EAD3",
    "Variables de retard":          "EAD1DC",
    "Variables de choc":            "FCE5CD",
    "Effets fixes / Saisonnalité":  "FFF2CC",
    "Variables de contrôle":        "CFE2F3",
    "Variables de revenu / profit": "D9D2E9",
}

current_row = 3
for row in rows:
    if row["__header"]:
        g = row["__groupe"]
        ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
        c = ws.cell(row=current_row, column=1, value=g)
        c.font = Font(name="Times New Roman", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E5090")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = Border(left=thick, right=thick, top=thick, bottom=thick)
        ws.row_dimensions[current_row].height = 16
        current_row += 1
    else:
        g = row["__groupe"]
        fill_color = GROUPE_COLORS.get(g, "FFFFFF")
        dataset_color = "E8F4E8" if row["Jeu de données"] == "E-Commerce" else "E8E8F4"
        for ci, key in enumerate(ENTETES, 1):
            val = row.get(key, "")
            c = ws.cell(row=current_row, column=ci, value=val)
            c.font = Font(name="Times New Roman", size=9)
            bg = dataset_color if ci == 2 else fill_color
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="left" if ci == 1 else "center",
                vertical="center"
            )
            c.border = border_data
        ws.row_dimensions[current_row].height = 14
        current_row += 1

note_row = current_row + 1
ws.cell(row=note_row, column=1,
        value="Note : N = nombre d'observations valides. Asymétrie > |1| indique une distribution asymétrique. Aplatissement > 3 indique une distribution leptokurtique.")
ws.cell(row=note_row, column=1).font = Font(name="Times New Roman", italic=True, size=8.5)
ws.cell(row=note_row + 1, column=1,
        value="Source : Auteur, à partir des données propriétaires E-Commerce et FMCG (2022–2024).")
ws.cell(row=note_row + 1, column=1).font = Font(name="Times New Roman", italic=True, size=8.5)

col_widths = [40, 16, 11, 11, 11, 11, 11, 8, 12, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

output_path = "Tableau_05_Statistiques_Descriptives.xlsx"
wb.save(output_path)
print(f"Fichier sauvegardé : {output_path}")
