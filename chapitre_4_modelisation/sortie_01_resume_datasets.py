import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df_ec = pd.read_excel("E-Ticaret_MMM_v2.xlsx")
df_fg = pd.read_excel("FMCG_MMM_v2.xlsx")

def get_info(df, nom):
    date_col = [c for c in df.columns if "Date" in c][0]
    dates = pd.to_datetime(df[date_col].dropna())
    freq_col = [c for c in df.columns if "Numéro de semaine" in c]
    canal_col = [c for c in df.columns if "Canal" in c]
    region_col = [c for c in df.columns if "Région" in c]
    sku_col = [c for c in df.columns if "SKU" in c]
    choc_cols = [c for c in df.columns if "Choc" in c]
    marketing_cols = [c for c in df.columns if c.strip().startswith("M :")]
    dep_var_col = [c for c in df.columns if "Commandes" in c or "Volume des ventes" in c and "original" not in c and "Retard" not in c and "Perte" not in c]

    return {
        "Jeu de données": nom,
        "Période couverte": f"{dates.min().strftime('%d/%m/%Y')} – {dates.max().strftime('%d/%m/%Y')}",
        "Fréquence temporelle": "Quotidienne" if "Canal" in " ".join(df.columns) else "Hebdomadaire",
        "Nombre d'observations": df.shape[0],
        "Nombre de variables": df.shape[1],
        "Unités d'analyse": f"{df[canal_col[0]].nunique()} canaux" if canal_col else f"{df[region_col[0]].nunique()} régions",
        "Nombre de SKU": df[sku_col[0]].nunique() if sku_col else "N/A",
        "Variable dépendante": dep_var_col[0].strip() if dep_var_col else "N/A",
        "Variables marketing (M :)": len(marketing_cols),
        "Variables de choc": len(choc_cols),
        "Valeurs manquantes": int(df.isnull().sum().sum()),
    }

infos = [get_info(df_ec, "E-Commerce"), get_info(df_fg, "FMCG")]

wb = Workbook()
ws = wb.active
ws.title = "Résumé des jeux de données"

titre = "Tableau 1 — Résumé des jeux de données"
ws.merge_cells("A1:D1")
ws["A1"] = titre
ws["A1"].font = Font(name="Times New Roman", bold=True, size=12)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 20

entetes = ["Caractéristique", "E-Commerce", "FMCG"]
couleur_entete = "1F3864"
for col_idx, val in enumerate(entetes, start=1):
    cell = ws.cell(row=2, column=col_idx, value=val)
    cell.font = Font(name="Times New Roman", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=couleur_entete)
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

cles = list(infos[0].keys())
cles.remove("Jeu de données")

couleur_paire = "DCE6F1"
thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, cle in enumerate(cles):
    row = i + 3
    fill_color = couleur_paire if i % 2 == 0 else "FFFFFF"
    c1 = ws.cell(row=row, column=1, value=cle)
    c1.font = Font(name="Times New Roman", size=10)
    c1.fill = PatternFill("solid", fgColor=fill_color)
    c1.alignment = Alignment(horizontal="left", vertical="center")
    c1.border = border

    for col_idx, info in enumerate(infos, start=2):
        c = ws.cell(row=row, column=col_idx, value=info[cle])
        c.font = Font(name="Times New Roman", size=10)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[row].height = 16

ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 26

source_row = len(cles) + 4
ws.cell(row=source_row, column=1, value="Source : Auteur, à partir des données propriétaires E-Commerce et FMCG (2022–2024).")
ws.cell(row=source_row, column=1).font = Font(name="Times New Roman", italic=True, size=9)

output_path = "Tableau_01_Resume_Jeux_de_Donnees.xlsx"
wb.save(output_path)
print(f"Fichier sauvegardé : {output_path}")
