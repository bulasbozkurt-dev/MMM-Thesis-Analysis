import pandas as pd
import numpy as np
import statsmodels.api as sm
from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.stats.stattools import durbin_watson
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Chargement des données ────────────────────────────────────────────────────
df_ec = pd.read_excel("E-Ticaret_MMM_v2.xlsx")
df_fg = pd.read_excel("FMCG_MMM_v2.xlsx")
df_ec.columns = [c.strip() for c in df_ec.columns]
df_fg.columns = [c.strip() for c in df_fg.columns]

VARS_EC = [
    "M : Google", "M : Meta", "M : Influenceur", "M : Indice de dépenses",
    "M : Retard t-1", "M : Retard t-7", "D : Retard t-1",
    "Choc : Taux de change", "Choc : Crise logistique",
    "Facteur saisonnier", "Effet du dimanche",
    "Dépenses des concurrents", "Part de voix (%)", "Taux de conversion (%)",
]
DEP_EC = "D : Commandes (canal)"

VARS_FG = [
    "M : Dépenses marketing totales", "M : Pondéré",
    "M : Retard 1 semaine", "M : Retard 4 semaines", "D : Retard 1 semaine",
    "Choc : Taux de change", "Choc : Impact du séisme",
    "Facteur saisonnier",
    "Dépenses marketing des concurrents", "Part de voix (%)", "Part de marché (%)",
    "Rupture de stock (%)", "Stock (jours)", "Efficacité marketing",
]
DEP_FG = "D : Volume des ventes"

def estimer_mco(df, dep, vars_list):
    valid = [v for v in vars_list if v in df.columns]
    sub   = df[[dep] + valid].dropna()
    Y     = sub[dep]
    X     = sm.add_constant(sub[valid], has_constant="add")
    mod   = sm.OLS(Y, X).fit()
    vif   = pd.DataFrame({
        "Variable": X.columns,
        "VIF": [variance_inflation_factor(X.values, i)
                for i in range(X.shape[1])]
    })
    return mod, vif

mod_ec, vif_ec = estimer_mco(df_ec, DEP_EC, VARS_EC)
mod_fg, vif_fg = estimer_mco(df_fg, DEP_FG, VARS_FG)

print("=== E-Commerce ===")
print(mod_ec.summary())
print("\n=== FMCG ===")
print(mod_fg.summary())

# ── Mise en forme Excel ───────────────────────────────────────────────────────
def sig(p):
    if p < 0.01: return "***"
    if p < 0.05: return "**"
    if p < 0.10: return "*"
    return ""

def build_rows(mod, vif_data):
    rows = []
    ci   = mod.conf_int()
    for name in mod.params.index:
        vif_row = vif_data[vif_data["Variable"] == name]
        vif_val = round(vif_row["VIF"].values[0], 2) if len(vif_row) > 0 else "—"
        rows.append({
            "Variable":      name,
            "Coefficient":   round(mod.params[name], 5),
            "Écart-type":    round(mod.bse[name], 5),
            "t de Student":  round(mod.tvalues[name], 3),
            "p-valeur":      round(mod.pvalues[name], 4),
            "Sig.":          sig(mod.pvalues[name]),
            "IC 95% inf.":   round(ci.loc[name, 0], 5),
            "IC 95% sup.":   round(ci.loc[name, 1], 5),
            "VIF":           vif_val,
        })
    return rows

rows_ec = build_rows(mod_ec, vif_ec)
rows_fg = build_rows(mod_fg, vif_fg)

ENTETES = ["Variable", "Coefficient", "Écart-type", "t de Student",
           "p-valeur", "Sig.", "IC 95% inf.", "IC 95% sup.", "VIF"]

GROUPE_MAP_EC = {
    "const":                        "Constante",
    "M : Google":                   "Variable explicative (Marketing)",
    "M : Meta":                     "Variable explicative (Marketing)",
    "M : Influenceur":              "Variable explicative (Marketing)",
    "M : Indice de dépenses":       "Variable explicative (Marketing)",
    "M : Retard t-1":               "Variable de retard",
    "M : Retard t-7":               "Variable de retard",
    "D : Retard t-1":               "Variable de retard",
    "Choc : Taux de change":        "Variable de choc",
    "Choc : Crise logistique":      "Variable de choc",
    "Facteur saisonnier":           "Effet fixe / Saisonnalité",
    "Effet du dimanche":            "Effet fixe / Saisonnalité",
    "Dépenses des concurrents":     "Variable de contrôle",
    "Part de voix (%)":             "Variable de contrôle",
    "Taux de conversion (%)":       "Variable de contrôle",
}
GROUPE_MAP_FG = {
    "const":                                "Constante",
    "M : Dépenses marketing totales":       "Variable explicative (Marketing)",
    "M : Pondéré":                          "Variable explicative (Marketing)",
    "M : Retard 1 semaine":                 "Variable de retard",
    "M : Retard 4 semaines":               "Variable de retard",
    "D : Retard 1 semaine":                "Variable de retard",
    "Choc : Taux de change":               "Variable de choc",
    "Choc : Impact du séisme":             "Variable de choc",
    "Facteur saisonnier":                  "Effet fixe / Saisonnalité",
    "Dépenses marketing des concurrents":  "Variable de contrôle",
    "Part de voix (%)":                    "Variable de contrôle",
    "Part de marché (%)":                  "Variable de contrôle",
    "Rupture de stock (%)":                "Variable de contrôle",
    "Stock (jours)":                       "Variable de contrôle",
    "Efficacité marketing":                "Variable de contrôle",
}
COULEURS_GROUPE = {
    "Constante":                        "EFEFEF",
    "Variable explicative (Marketing)": "D9EAD3",
    "Variable de retard":               "EAD1DC",
    "Variable de choc":                 "FCE5CD",
    "Effet fixe / Saisonnalité":        "FFF2CC",
    "Variable de contrôle":             "CFE2F3",
}

def couleur_sig(s):
    if s == "***": return "1E6B3C"
    if s == "**":  return "1F3864"
    if s == "*":   return "B45309"
    return "888888"

thin  = Side(style="thin",   color="AAAAAA")
thick = Side(style="medium", color="1F3864")
b     = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def ecrire_resultats(wb, rows, groupe_map, titre, nom_feuille, mod, premier):
    ws = wb.active if premier else wb.create_sheet(nom_feuille)
    if premier:
        ws.title = nom_feuille

    N    = len(ENTETES)
    last = get_column_letter(N)

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = titre
    ws["A1"].font      = Font(name="Times New Roman", bold=True, size=11)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(ENTETES, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = Font(name="Times New Roman", bold=True, color="FFFFFF", size=9.5)
        c.fill      = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = b
    ws.row_dimensions[2].height = 26

    prev_groupe = None
    cur = 3
    for i, row in enumerate(rows):
        var    = row["Variable"]
        groupe = groupe_map.get(var, "Autre")
        fill_hex = COULEURS_GROUPE.get(groupe, "FFFFFF")
        bg = fill_hex if i % 2 == 0 else "FFFFFF"

        if groupe != prev_groupe:
            ws.merge_cells(f"A{cur}:{last}{cur}")
            gc = ws.cell(row=cur, column=1, value=groupe)
            gc.font      = Font(name="Times New Roman", bold=True, size=9, color="FFFFFF")
            gc.fill      = PatternFill("solid", fgColor="2E5090")
            gc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            gc.border    = b
            ws.row_dimensions[cur].height = 15
            cur += 1
            prev_groupe = groupe

        for ci, key in enumerate(ENTETES, 1):
            val = row.get(key, "")
            c   = ws.cell(row=cur, column=ci, value=val)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="left" if ci == 1 else "center",
                vertical="center"
            )
            c.border = b
            if key == "Sig." and val:
                c.font = Font(name="Times New Roman", size=10, bold=True,
                              color=couleur_sig(val))
            elif key == "Coefficient" and row.get("Sig.") in ("***", "**"):
                c.font = Font(name="Times New Roman", size=9, bold=True)
            else:
                c.font = Font(name="Times New Roman", size=9)
        ws.row_dimensions[cur].height = 14
        cur += 1

    # ── Indicateurs globaux ───────────────────────────────────────────────────
    cur += 1
    metriques = [
        ("R²",               round(mod.rsquared, 4)),
        ("R² ajusté",        round(mod.rsquared_adj, 4)),
        ("Statistique F",    round(mod.fvalue, 3)),
        ("p-valeur (F)",     f"{mod.f_pvalue:.4e}"),
        ("AIC",              round(mod.aic, 2)),
        ("BIC",              round(mod.bic, 2)),
        ("Durbin-Watson",    round(durbin_watson(mod.resid), 3)),
        ("Log-vraisemblance",round(mod.llf, 2)),
        ("Nb. observations", int(mod.nobs)),
        ("Nb. variables",    int(mod.df_model + 1)),
    ]

    ws.merge_cells(f"A{cur}:{last}{cur}")
    mc = ws.cell(row=cur, column=1, value="Indicateurs d'ajustement global")
    mc.font      = Font(name="Times New Roman", bold=True, size=10, color="FFFFFF")
    mc.fill      = PatternFill("solid", fgColor="1F3864")
    mc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    mc.border    = b
    ws.row_dimensions[cur].height = 17
    cur += 1

    for j, (label, val) in enumerate(metriques):
        bg = "DCE6F1" if j % 2 == 0 else "FFFFFF"
        c1 = ws.cell(row=cur, column=1, value=label)
        c1.font = Font(name="Times New Roman", size=9, bold=True)
        c1.fill = PatternFill("solid", fgColor=bg)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c1.border = b
        c2 = ws.cell(row=cur, column=2, value=val)
        c2.font = Font(name="Times New Roman", size=9)
        c2.fill = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = b
        for ci in range(3, N+1):
            cx = ws.cell(row=cur, column=ci)
            cx.fill   = PatternFill("solid", fgColor=bg)
            cx.border = b
        ws.row_dimensions[cur].height = 14
        cur += 1

    note = cur + 1
    ws.cell(row=note, column=1,
            value="Note : *** p < 0,01 ; ** p < 0,05 ; * p < 0,10. "
                  "IC 95% = intervalle de confiance. VIF = Facteur d'inflation de la variance.")
    ws.cell(row=note, column=1).font = Font(name="Times New Roman", italic=True, size=8.5)
    ws.cell(row=note+1, column=1,
            value="Source : Auteur, estimations MCO (statsmodels). Données propriétaires (2022–2024).")
    ws.cell(row=note+1, column=1).font = Font(name="Times New Roman", italic=True, size=8.5)

    largeurs = [34, 13, 13, 13, 11, 7, 13, 13, 9]
    for i, w in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

ecrire_resultats(wb, rows_ec, GROUPE_MAP_EC,
                 "Tableau 12a — Résultats MCO complets — E-Commerce",
                 "MCO E-Commerce", mod_ec, premier=True)
ecrire_resultats(wb, rows_fg, GROUPE_MAP_FG,
                 "Tableau 12b — Résultats MCO complets — FMCG",
                 "MCO FMCG", mod_fg, premier=False)

wb.save("Tableau_12_Resultats_MCO_Complets.xlsx")
print("Fichier sauvegardé : Tableau_12_Resultats_MCO_Complets.xlsx")
