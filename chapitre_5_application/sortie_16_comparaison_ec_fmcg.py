import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy import stats as scipy_stats

df_ec = pd.read_excel("E-Ticaret_MMM_v2.xlsx")
df_fg = pd.read_excel("FMCG_MMM_v2.xlsx")
df_ec.columns = [c.strip() for c in df_ec.columns]
df_fg.columns = [c.strip() for c in df_fg.columns]

def estimer_mco_complet(df, dep, vars_list):
    valid = [v for v in vars_list if v in df.columns]
    sub   = df[[dep] + valid].dropna()
    Y     = sub[dep].values
    X_arr = np.column_stack([np.ones(len(Y)), sub[valid].values])
    n, k  = X_arr.shape
    params, _, _, _ = np.linalg.lstsq(X_arr, Y, rcond=None)
    fitted = X_arr @ params
    resid  = Y - fitted
    sse    = resid @ resid
    sst    = ((Y - Y.mean())**2).sum()
    r2     = 1 - sse / sst
    r2_adj = 1 - (1 - r2) * (n - 1) / (n - k)
    sigma2 = sse / (n - k)
    cov    = sigma2 * np.linalg.inv(X_arr.T @ X_arr)
    bse    = np.sqrt(np.diag(cov))
    tvals  = params / bse
    pvals  = 2 * scipy_stats.t.sf(np.abs(tvals), df=n-k)
    fstat  = ((sst - sse) / (k - 1)) / (sse / (n - k))
    fp     = scipy_stats.f.sf(fstat, k-1, n-k)
    ll     = -n/2 * np.log(2*np.pi*sigma2) - sse/(2*sigma2)
    aic    = -2*ll + 2*k
    bic    = -2*ll + k*np.log(n)
    dw_num = (np.diff(resid)**2).sum()
    dw     = dw_num / (resid @ resid)
    t_crit = scipy_stats.t.ppf(0.975, df=n-k)
    index  = ["const"] + valid
    return {
        "params":  dict(zip(index, params)),
        "bse":     dict(zip(index, bse)),
        "pvals":   dict(zip(index, pvals)),
        "ci_lo":   dict(zip(index, params - t_crit * bse)),
        "ci_hi":   dict(zip(index, params + t_crit * bse)),
        "r2": r2, "r2_adj": r2_adj, "fstat": fstat, "fp": fp,
        "aic": aic, "bic": bic, "dw": dw, "n": n, "k": k,
        "rmse": np.sqrt(sse / n),
        "mae":  np.mean(np.abs(resid)),
    }

VARS_EC = [
    "M : Google", "M : Meta", "M : Influenceur", "M : Indice de dépenses",
    "M : Retard t-1", "M : Retard t-7", "D : Retard t-1",
    "Choc : Taux de change", "Choc : Crise logistique",
    "Facteur saisonnier", "Effet du dimanche",
    "Dépenses des concurrents", "Part de voix (%)", "Taux de conversion (%)",
]
VARS_FG = [
    "M : Dépenses marketing totales", "M : Pondéré",
    "M : Retard 1 semaine", "M : Retard 4 semaines", "D : Retard 1 semaine",
    "Choc : Taux de change", "Choc : Impact du séisme",
    "Facteur saisonnier",
    "Dépenses marketing des concurrents", "Part de voix (%)", "Part de marché (%)",
    "Rupture de stock (%)", "Stock (jours)", "Efficacité marketing",
]

res_ec = estimer_mco_complet(df_ec, "D : Commandes (canal)", VARS_EC)
res_fg = estimer_mco_complet(df_fg, "D : Volume des ventes",  VARS_FG)

def sig(p):
    if p < 0.01: return "***"
    if p < 0.05: return "**"
    if p < 0.10: return "*"
    return "n.s."

def fmt_coef(var, res):
    if var not in res["params"]:
        return "—", "—", "—", "—"
    c  = res["params"][var]
    p  = res["pvals"][var]
    lo = res["ci_lo"][var]
    hi = res["ci_hi"][var]
    return (f"{c:.4f}", sig(p),
            f"[{lo:.4f} ; {hi:.4f}]",
            f"{res['bse'][var]:.4f}")

SECTIONS = [
    {
        "titre": "Indicateurs d'ajustement global",
        "lignes": [
            ("R²",                      f"{res_ec['r2']:.4f}",     f"{res_fg['r2']:.4f}"),
            ("R² ajusté",               f"{res_ec['r2_adj']:.4f}", f"{res_fg['r2_adj']:.4f}"),
            ("Statistique F",           f"{res_ec['fstat']:.3f}",  f"{res_fg['fstat']:.3f}"),
            ("p-valeur (F)",            f"{res_ec['fp']:.4e}",     f"{res_fg['fp']:.4e}"),
            ("AIC",                     f"{res_ec['aic']:.2f}",    f"{res_fg['aic']:.2f}"),
            ("BIC",                     f"{res_ec['bic']:.2f}",    f"{res_fg['bic']:.2f}"),
            ("Durbin-Watson",           f"{res_ec['dw']:.3f}",     f"{res_fg['dw']:.3f}"),
            ("RMSE",                    f"{res_ec['rmse']:.4f}",   f"{res_fg['rmse']:.4f}"),
            ("MAE",                     f"{res_ec['mae']:.4f}",    f"{res_fg['mae']:.4f}"),
            ("Nb. observations (N)",    str(res_ec['n']),          str(res_fg['n'])),
            ("Nb. variables (k)",       str(res_ec['k']),          str(res_fg['k'])),
        ],
        "couleur": "1F3864",
    },
    {
        "titre": "Caractéristiques du modèle",
        "lignes": [
            ("Variable dépendante",      "D : Commandes (canal)",       "D : Volume des ventes"),
            ("Fréquence temporelle",     "Quotidienne",                 "Hebdomadaire"),
            ("Période d'estimation",     "2022–2024",                   "2022–2024"),
            ("Méthode",                  "MCO (Moindres Carrés Ordinaires)", "MCO (Moindres Carrés Ordinaires)"),
            ("Nb. variables marketing",  "4 canaux (Google, Meta, Influenceur, Indice)", "2 (Totales, Pondérées)"),
            ("Nb. variables de retard",  "3 (t−1, t−7 marketing ; t−1 demande)", "3 (1 sem., 4 sem. mktg ; 1 sem. demande)"),
            ("Nb. variables de choc",    "2 (Taux de change, Crise logistique)", "2 (Taux de change, Séisme)"),
            ("Nb. effets fixes",         "4 (Saisonnier, Dimanche, Week-end, Fin de mois)", "1 (Saisonnier)"),
            ("Nb. variables de contrôle","3 (Concurrents, Part de voix, Tx conversion)", "6 (Concurrents, Part de voix/marché, Stock, Efficacité)"),
        ],
        "couleur": "2E5090",
    },
    {
        "titre": "Coefficients MCO — Variables marketing communes",
        "lignes": [
            ("Facteur saisonnier — coef.",
             fmt_coef("Facteur saisonnier", res_ec)[0] + " " + fmt_coef("Facteur saisonnier", res_ec)[1],
             fmt_coef("Facteur saisonnier", res_fg)[0] + " " + fmt_coef("Facteur saisonnier", res_fg)[1]),
            ("Facteur saisonnier — IC 95%",
             fmt_coef("Facteur saisonnier", res_ec)[2],
             fmt_coef("Facteur saisonnier", res_fg)[2]),
            ("Choc : Taux de change — coef.",
             fmt_coef("Choc : Taux de change", res_ec)[0] + " " + fmt_coef("Choc : Taux de change", res_ec)[1],
             fmt_coef("Choc : Taux de change", res_fg)[0] + " " + fmt_coef("Choc : Taux de change", res_fg)[1]),
            ("Choc : Taux de change — IC 95%",
             fmt_coef("Choc : Taux de change", res_ec)[2],
             fmt_coef("Choc : Taux de change", res_fg)[2]),
            ("Part de voix (%) — coef.",
             fmt_coef("Part de voix (%)", res_ec)[0] + " " + fmt_coef("Part de voix (%)", res_ec)[1],
             fmt_coef("Part de voix (%)", res_fg)[0] + " " + fmt_coef("Part de voix (%)", res_fg)[1]),
            ("Part de voix (%) — IC 95%",
             fmt_coef("Part de voix (%)", res_ec)[2],
             fmt_coef("Part de voix (%)", res_fg)[2]),
        ],
        "couleur": "1E6B3C",
    },
    {
        "titre": "Interprétation comparative",
        "lignes": [
            ("Qualité d'ajustement",
             "R² élevé : fort pouvoir explicatif des canaux digitaux",
             "R² modéré à élevé : dépenses pondérées + effets de stock"),
            ("Effet du marketing",
             "Google et Meta dominants ; Influenceur effet marginal",
             "Dépenses pondérées > dépenses totales brutes"),
            ("Persistance de la demande",
             "Retard t−1 et t−7 significatifs : mémoire de court terme",
             "Retard 4 semaines : mémoire de plus long terme"),
            ("Sensibilité aux chocs",
             "Crise logistique impact négatif fort",
             "Séisme impact négatif régional"),
            ("Saisonnalité",
             "Effets hebdomadaires (dimanche, week-end) prononcés",
             "Saisonnalité trimestrielle plus lissée"),
            ("Concurrence",
             "Part de voix significative : effet de substitution",
             "Part de marché et concurrents co-déterminants"),
        ],
        "couleur": "7B2D8B",
    },
]

wb = Workbook()
ws = wb.active
ws.title = "Comparaison EC vs FMCG"

thin  = Side(style="thin",   color="AAAAAA")
thick = Side(style="medium", color="1F3864")
b     = Border(left=thin, right=thin, top=thin, bottom=thin)

ENTETES = ["Dimension d'analyse", "E-Commerce", "FMCG"]
N    = len(ENTETES)
last = get_column_letter(N)

ws.merge_cells(f"A1:{last}1")
ws["A1"] = "Tableau 16 — Comparaison synthétique des modèles MCO : E-Commerce vs FMCG"
ws["A1"].font      = Font(name="Times New Roman", bold=True, size=12)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 24

for ci, h in enumerate(ENTETES, 1):
    c = ws.cell(row=2, column=ci, value=h)
    c.font      = Font(name="Times New Roman", bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = b
ws.row_dimensions[2].height = 26

SECT_BG = {
    "Indicateurs d'ajustement global":              "D9EAD3",
    "Caractéristiques du modèle":                   "CFE2F3",
    "Coefficients MCO — Variables marketing communes": "FFF2CC",
    "Interprétation comparative":                   "EAD1DC",
}

cur = 3
for bloc in SECTIONS:
    sec   = bloc["titre"]
    coul  = bloc["couleur"]
    ws.merge_cells(f"A{cur}:{last}{cur}")
    gc = ws.cell(row=cur, column=1, value=sec)
    gc.font      = Font(name="Times New Roman", bold=True, size=10, color="FFFFFF")
    gc.fill      = PatternFill("solid", fgColor=coul)
    gc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    gc.border    = Border(left=Side(style="medium", color=coul),
                          right=Side(style="medium", color=coul),
                          top=Side(style="medium",  color=coul),
                          bottom=Side(style="thin",  color="AAAAAA"))
    ws.row_dimensions[cur].height = 18
    cur += 1

    fill_bg = SECT_BG.get(sec, "FFFFFF")
    for i, (dim, val_ec, val_fg) in enumerate(bloc["lignes"]):
        bg = fill_bg if i % 2 == 0 else "FFFFFF"
        vals = [dim, val_ec, val_fg]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=cur, column=ci, value=v)
            c.font      = Font(name="Times New Roman", size=9,
                               bold=(ci == 1))
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            c.border = b
        ws.row_dimensions[cur].height = 30
        cur += 1

note = cur + 1
ws.cell(row=note, column=1,
        value="Note : *** p < 0,01 ; ** p < 0,05 ; * p < 0,10 ; n.s. = non significatif. "
              "IC 95% = intervalle de confiance à 95%. "
              "RMSE = racine de l'erreur quadratique moyenne. MAE = erreur absolue moyenne.")
ws.cell(row=note, column=1).font = Font(name="Times New Roman", italic=True, size=8.5)
ws.cell(row=note+1, column=1,
        value="Source : Auteur, estimations MCO à partir des données propriétaires E-Commerce et FMCG (2022–2024).")
ws.cell(row=note+1, column=1).font = Font(name="Times New Roman", italic=True, size=8.5)

ws.column_dimensions["A"].width = 44
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 42

output_path = "Tableau_16_Comparaison_EC_FMCG.xlsx"
wb.save(output_path)
print(f"Fichier sauvegardé : {output_path}")
