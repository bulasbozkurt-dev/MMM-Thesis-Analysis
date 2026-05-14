import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy import stats as scipy_stats

df_ec = pd.read_excel("E-Ticaret_MMM_v2.xlsx")
df_fg = pd.read_excel("FMCG_MMM_v2.xlsx")
df_ec.columns = [c.strip() for c in df_ec.columns]
df_fg.columns = [c.strip() for c in df_fg.columns]

def estimer(df, dep, vars_list):
    valid = [v for v in vars_list if v in df.columns]
    if not valid:
        return None
    sub   = df[[dep] + valid].dropna()
    Y     = sub[dep].values
    X     = np.column_stack([np.ones(len(Y)), sub[valid].values])
    n, k  = X.shape
    params, _, _, _ = np.linalg.lstsq(X, Y, rcond=None)
    fitted = X @ params
    resid  = Y - fitted
    sse    = resid @ resid
    sst    = ((Y - Y.mean())**2).sum()
    r2     = 1 - sse / sst
    r2_adj = 1 - (1 - r2) * (n-1) / (n-k)
    sigma2 = sse / (n - k)
    try:
        cov   = sigma2 * np.linalg.inv(X.T @ X)
        bse   = np.sqrt(np.diag(cov))
    except Exception:
        bse   = np.full(k, np.nan)
    tvals  = params / bse
    pvals  = 2 * scipy_stats.t.sf(np.abs(tvals), df=n-k)
    fstat  = ((sst - sse) / (k-1)) / (sse / (n-k))
    fp     = scipy_stats.f.sf(fstat, k-1, n-k)
    ll     = -n/2 * np.log(2*np.pi*sigma2) - sse/(2*sigma2)
    aic    = -2*ll + 2*k
    bic    = -2*ll + k*np.log(n)
    dw     = (np.diff(resid)**2).sum() / (resid @ resid)
    rmse   = np.sqrt(sse / n)
    index  = ["const"] + valid
    t_crit = scipy_stats.t.ppf(0.975, df=n-k)
    return {
        "params": dict(zip(index, params)),
        "bse":    dict(zip(index, bse)),
        "pvals":  dict(zip(index, pvals)),
        "ci_lo":  dict(zip(index, params - t_crit * bse)),
        "ci_hi":  dict(zip(index, params + t_crit * bse)),
        "r2": r2, "r2_adj": r2_adj, "fstat": fstat, "fp": fp,
        "aic": aic, "bic": bic, "dw": dw, "rmse": rmse,
        "n": n, "k": k,
    }

def sig(p):
    if pd.isna(p): return "—"
    if p < 0.01:   return "***"
    if p < 0.05:   return "**"
    if p < 0.10:   return "*"
    return "n.s."

BASE_EC = [
    "M : Google", "M : Meta", "M : Influenceur", "M : Indice de dépenses",
    "M : Retard t-1", "M : Retard t-7", "D : Retard t-1",
    "Choc : Taux de change", "Choc : Crise logistique",
    "Facteur saisonnier", "Effet du dimanche",
    "Dépenses des concurrents", "Part de voix (%)", "Taux de conversion (%)",
]
VARIANTES_EC = {
    "Modèle complet\n(référence)":         BASE_EC,
    "Sans retards":                        [v for v in BASE_EC if "Retard" not in v],
    "Sans effets fixes":                   [v for v in BASE_EC if v not in
                                            ["Facteur saisonnier", "Effet du dimanche"]],
    "Sans chocs":                          [v for v in BASE_EC if "Choc" not in v],
    "Marketing\nuniquement":               [v for v in BASE_EC if v.startswith("M :")],
    "Marketing\n+ retards":                [v for v in BASE_EC
                                            if v.startswith("M :") or "Retard" in v],
}

BASE_FG = [
    "M : Dépenses marketing totales", "M : Pondéré",
    "M : Retard 1 semaine", "M : Retard 4 semaines", "D : Retard 1 semaine",
    "Choc : Taux de change", "Choc : Impact du séisme",
    "Facteur saisonnier",
    "Dépenses marketing des concurrents", "Part de voix (%)", "Part de marché (%)",
    "Rupture de stock (%)", "Stock (jours)", "Efficacité marketing",
]
VARIANTES_FG = {
    "Modèle complet\n(référence)":         BASE_FG,
    "Sans retards":                        [v for v in BASE_FG if "Retard" not in v],
    "Sans effets de stock":                [v for v in BASE_FG if v not in
                                            ["Rupture de stock (%)", "Stock (jours)"]],
    "Sans chocs":                          [v for v in BASE_FG if "Choc" not in v],
    "Marketing\nuniquement":               [v for v in BASE_FG if v.startswith("M :")],
    "Marketing\n+ retards":                [v for v in BASE_FG
                                            if v.startswith("M :") or "Retard" in v],
}

METRIQUES  = ["r2", "r2_adj", "fstat", "fp", "aic", "bic", "dw", "rmse", "n", "k"]
LABELS_M   = ["R²", "R² ajusté", "Stat. F", "p-val. (F)",
              "AIC", "BIC", "Durbin-Watson", "RMSE", "N obs.", "k vars."]
FILL_METR  = {
    "r2":     "D9EAD3", "r2_adj": "D9EAD3",
    "fstat":  "CFE2F3", "fp":     "CFE2F3",
    "aic":    "FFF2CC", "bic":    "FFF2CC",
    "dw":     "EAD1DC", "rmse":   "F4CCCC",
    "n":      "EFEFEF", "k":      "EFEFEF",
}

def fmt(key, val):
    if val is None: return "—"
    if key in ("n", "k"):          return str(int(val))
    if key == "fp":                return f"{val:.4e}"
    if key in ("r2", "r2_adj", "dw"): return f"{val:.4f}"
    if key == "fstat":             return f"{val:.3f}"
    return f"{val:.2f}"

thin  = Side(style="thin",   color="AAAAAA")
thick = Side(style="medium", color="1F3864")
b     = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def ecrire_feuille(wb, variantes, df_data, dep,
                   titre, nom_feuille, mark_prefix, premier):
    ws = wb.active if premier else wb.create_sheet(nom_feuille)
    if premier:
        ws.title = nom_feuille

    resultats = {nom: estimer(df_data, dep, vlist)
                 for nom, vlist in variantes.items()}
    noms   = list(variantes.keys())
    N      = len(noms) + 1
    last   = get_column_letter(N)

    # titre
    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = titre
    ws["A1"].font      = Font(name="Times New Roman", bold=True, size=11)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # en-têtes colonnes
    ws.cell(row=2, column=1, value="Indicateur").font = Font(
        name="Times New Roman", bold=True, color="FFFFFF", size=9.5)
    ws.cell(row=2, column=1).fill      = PatternFill("solid", fgColor="1F3864")
    ws.cell(row=2, column=1).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=2, column=1).border = b

    for ci, nom in enumerate(noms, 2):
        c = ws.cell(row=2, column=ci, value=nom.replace("\n", " "))
        c.font      = Font(name="Times New Roman", bold=True,
                           color="FFFFFF", size=8.5)
        c.fill      = PatternFill("solid",
                                  fgColor="1F3864" if "référence" in nom else "2E5090")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = b
    ws.row_dimensions[2].height = 44

    # métriques
    for ri, (key, label) in enumerate(zip(METRIQUES, LABELS_M)):
        cur = ri + 3
        bg  = FILL_METR.get(key, "FFFFFF")
        c1  = ws.cell(row=cur, column=1, value=label)
        c1.font      = Font(name="Times New Roman", size=9, bold=True)
        c1.fill      = PatternFill("solid", fgColor=bg)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c1.border    = b
        for ci, nom in enumerate(noms, 2):
            res = resultats[nom]
            val = res[key] if res else None
            c   = ws.cell(row=cur, column=ci, value=fmt(key, val))
            c.font      = Font(name="Times New Roman", size=9,
                               bold=("référence" in nom))
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = b
        ws.row_dimensions[cur].height = 16

    # section coefficients marketing
    mark_vars = [v for v in list(variantes.values())[0]
                 if v.startswith(mark_prefix)]
    sep = len(METRIQUES) + 4
    ws.merge_cells(f"A{sep}:{last}{sep}")
    gc = ws.cell(row=sep, column=1,
                 value="Coefficients MCO — Variables marketing")
    gc.font      = Font(name="Times New Roman", bold=True,
                        size=10, color="FFFFFF")
    gc.fill      = PatternFill("solid", fgColor="1E6B3C")
    gc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    gc.border    = b
    ws.row_dimensions[sep].height = 18

    for ri2, var in enumerate(mark_vars):
        cur2 = sep + 1 + ri2
        bg   = "D9EAD3" if ri2 % 2 == 0 else "FFFFFF"
        c1   = ws.cell(row=cur2, column=1, value=var)
        c1.font      = Font(name="Times New Roman", size=9)
        c1.fill      = PatternFill("solid", fgColor=bg)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c1.border    = b
        for ci, nom in enumerate(noms, 2):
            res = resultats[nom]
            if res and var in res["params"]:
                val = f"{res['params'][var]:.4f} {sig(res['pvals'][var])}"
            else:
                val = "—"
            c   = ws.cell(row=cur2, column=ci, value=val)
            c.font      = Font(name="Times New Roman", size=9)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = b
        ws.row_dimensions[cur2].height = 14

    note = sep + len(mark_vars) + 2
    ws.cell(row=note, column=1,
            value="Note : *** p < 0,01 ; ** p < 0,05 ; * p < 0,10 ; n.s. = non significatif. "
                  "Le modèle complet est la spécification de référence.")
    ws.cell(row=note,   column=1).font = Font(name="Times New Roman",
                                               italic=True, size=8.5)
    ws.cell(row=note+1, column=1,
            value="Source : Auteur, estimations MCO à partir des données propriétaires (2022–2024).")
    ws.cell(row=note+1, column=1).font = Font(name="Times New Roman",
                                               italic=True, size=8.5)

    ws.column_dimensions["A"].width = 20
    for ci in range(2, N+1):
        ws.column_dimensions[get_column_letter(ci)].width = 20

ecrire_feuille(
    wb, VARIANTES_EC, df_ec, "D : Commandes (canal)",
    "Appendice D1 — Comparaison des variantes MCO — E-Commerce",
    "Variantes E-Commerce", "M :", premier=True
)
ecrire_feuille(
    wb, VARIANTES_FG, df_fg, "D : Volume des ventes",
    "Appendice D2 — Comparaison des variantes MCO — FMCG",
    "Variantes FMCG", "M :", premier=False
)

output_path = "Appendice_D_Variantes_Modeles.xlsx"
wb.save(output_path)
print(f"Fichier sauvegardé : {output_path}")
